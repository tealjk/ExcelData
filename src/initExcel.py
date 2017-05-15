#!/usr/bin/python


from openpyxl import load_workbook

from openpyxl import Workbook

wb2 = load_workbook('/Users/junkim/Desktop/PythonTOexcel/staging/testDataInput/testdata.xlsx')

print wb2.get_sheet_names()





# wb = Workbook()
#ws = wb.active
#ws1 = wb.create_sheet("Mysheet")
#ws.title = "NewTitle"






